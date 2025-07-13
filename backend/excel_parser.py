import pandas as pd
import openpyxl
from typing import Dict, List, Any
import re
import os

class ExcelParser:
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm']
        
    def parse_workbook(self, excel_path: str) -> Dict[str, Any]:
        """Parse Excel workbook and extract all data"""
        try:
            file_extension = os.path.splitext(excel_path)[1].lower()
            if file_extension not in self.supported_extensions:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Read all sheets
            sheets_data = {}
            
            if file_extension == '.xlsx' or file_extension == '.xlsm':
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_data = self._parse_sheet_openpyxl(sheet, sheet_name)
                    if sheet_data["numbers"]:  # Only include sheets with numbers
                        sheets_data[sheet_name] = sheet_data
                        
            else:  # .xls files
                xl_file = pd.ExcelFile(excel_path)
                for sheet_name in xl_file.sheet_names:
                    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                    sheet_data = self._parse_sheet_pandas(df, sheet_name)
                    if sheet_data["numbers"]:
                        sheets_data[sheet_name] = sheet_data
            
            print(f"Parsed {len(sheets_data)} sheets from {os.path.basename(excel_path)}")
            return {
                "filename": os.path.basename(excel_path),
                "sheets": sheets_data,
                "total_numbers": sum(len(sheet["numbers"]) for sheet in sheets_data.values())
            }
            
        except Exception as e:
            print(f"Error parsing Excel file {excel_path}: {str(e)}")
            raise Exception(f"Failed to parse Excel file: {str(e)}")
    
    def _parse_sheet_openpyxl(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Parse sheet using openpyxl"""
        numbers = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_info = self._extract_cell_info(cell, sheet_name)
                    if cell_info:
                        numbers.append(cell_info)
        
        return {
            "sheet_name": sheet_name,
            "numbers": numbers,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column
        }
    
    def _parse_sheet_pandas(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Parse sheet using pandas"""
        numbers = []
        
        for row_idx, row in df.iterrows():
            for col_idx, value in row.items():
                if pd.notna(value) and isinstance(value, (int, float)):
                    cell_ref = f"{self._col_num_to_letter(col_idx + 1)}{row_idx + 1}"
                    
                    # Get context from nearby cells
                    context = self._get_context_pandas(df, row_idx, col_idx)
                    
                    numbers.append({
                        "value": float(value),
                        "cell_reference": cell_ref,
                        "row": row_idx + 1,
                        "column": col_idx + 1,
                        "sheet_name": sheet_name,
                        "context": context,
                        "data_type": "number"
                    })
        
        return {
            "sheet_name": sheet_name,
            "numbers": numbers,
            "max_row": len(df),
            "max_column": len(df.columns)
        }
    
    def _extract_cell_info(self, cell, sheet_name: str) -> Dict[str, Any]:
        """Extract information from a cell"""
        try:
            value = cell.value
            
            if isinstance(value, (int, float)) and value != 0:
                # Get nearby cell values for context
                context = self._get_context_openpyxl(cell)
                
                return {
                    "value": float(value),
                    "cell_reference": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "sheet_name": sheet_name,
                    "context": context,
                    "data_type": "number"
                }
            
            elif isinstance(value, str) and value.strip():
                # Check if string contains numbers
                numbers_in_text = self._extract_numbers_from_text(value)
                if numbers_in_text:
                    context = self._get_context_openpyxl(cell)
                    return {
                        "value": numbers_in_text[0],  # Take first number found
                        "cell_reference": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "sheet_name": sheet_name,
                        "context": context,
                        "data_type": "text_with_number",
                        "original_text": value
                    }
            
            return None
            
        except Exception as e:
            print(f"Error extracting cell info: {str(e)}")
            return None
    
    def _get_context_openpyxl(self, cell) -> str:
        """Get context from nearby cells (headers, labels)"""
        context_parts = []
        sheet = cell.parent
        
        # Check cell to the left (potential row header)
        if cell.column > 1:
            left_cell = sheet.cell(row=cell.row, column=cell.column - 1)
            if left_cell.value and isinstance(left_cell.value, str):
                context_parts.append(left_cell.value.strip())
        
        # Check cell above (potential column header)
        if cell.row > 1:
            above_cell = sheet.cell(row=cell.row - 1, column=cell.column)
            if above_cell.value and isinstance(above_cell.value, str):
                context_parts.append(above_cell.value.strip())
        
        # Check top-left for potential section header
        if cell.row > 1 and cell.column > 1:
            topleft_cell = sheet.cell(row=cell.row - 1, column=cell.column - 1)
            if topleft_cell.value and isinstance(topleft_cell.value, str):
                context_parts.append(topleft_cell.value.strip())
        
        return " | ".join(context_parts) if context_parts else ""
    
    def _get_context_pandas(self, df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
        """Get context from nearby cells using pandas"""
        context_parts = []
        
        # Check cell to the left
        if col_idx > 0:
            left_value = df.iloc[row_idx, col_idx - 1]
            if pd.notna(left_value) and isinstance(left_value, str):
                context_parts.append(str(left_value).strip())
        
        # Check cell above
        if row_idx > 0:
            above_value = df.iloc[row_idx - 1, col_idx]
            if pd.notna(above_value) and isinstance(above_value, str):
                context_parts.append(str(above_value).strip())
        
        return " | ".join(context_parts) if context_parts else ""
    
    def _extract_numbers_from_text(self, text: str) -> List[float]:
        """Extract numbers from text string"""
        # Pattern to match numbers with various formats
        pattern = r'[\d,]+\.?\d*'
        matches = re.findall(pattern, text)
        
        numbers = []
        for match in matches:
            try:
                clean_number = match.replace(',', '')
                numbers.append(float(clean_number))
            except ValueError:
                continue
        
        return numbers
    
    def _col_num_to_letter(self, col_num: int) -> str:
        """Convert column number to Excel letter format"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result